#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Wolf Pack Dashboard v5 — Data Generator
在 Mac Mini 上執行，讀取 ETF 歷史數據，產出前端需要的 JSON 檔案。

用法:
  python3 generate_dashboard_data.py

輸出:
  ../data/dashboard.json  (主要儀表板數據)
  ../data/etf_pages.json  (各 ETF 個別頁面數據)
"""

import sys, json, os, glob, platform
from pathlib import Path
from datetime import datetime
import pandas as pd
import numpy as np

# ═══════════════════════════════════════
# 路徑設定（自動偵測 Mac / Windows）
# ═══════════════════════════════════════
if platform.system() == "Windows":
    _gdrive = Path("G:/其他電腦/我的 Mac/FinanceData/history/ETF")
    BASE = _gdrive if _gdrive.exists() else Path(os.path.expanduser("~/FinanceData/history/ETF"))
else:
    BASE = Path(os.path.expanduser("~/FinanceData/history/ETF"))
SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR.parent / "data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def fetch_market_indices(start_year=2025, start_month=10):
    """自動抓取加權指數 (TAIEX) 和櫃買指數 (TPEX) 歷史資料"""
    import requests, re

    taiex = {}
    tpex = {}
    now = datetime.now()

    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'}

    # 遍歷每個月
    y, m = start_year, start_month
    while True:
        if y > now.year or (y == now.year and m > now.month):
            break

        # ── TAIEX (TWSE) ──
        try:
            url = f'https://www.twse.com.tw/exchangeReport/FMTQIK?response=json&date={y}{m:02d}01'
            r = requests.get(url, headers=headers, timeout=10)
            data = r.json()
            if data.get('stat') == 'OK' and data.get('data'):
                for row in data['data']:
                    # row[0] = '115/03/13', row[4] = '33,400.32'
                    parts = row[0].split('/')
                    if len(parts) == 3:
                        dt = f"{int(parts[0]) + 1911}-{parts[1]}-{parts[2]}"
                        val = row[4].replace(',', '')
                        try:
                            taiex[dt] = float(val)
                        except ValueError:
                            pass
        except Exception as e:
            print(f"  ⚠️ TAIEX {y}/{m:02d} 抓取失敗: {e}")

        # ── TPEX (櫃買) ──
        try:
            roc_y = y - 1911
            url = f'https://www.tpex.org.tw/web/stock/aftertrading/daily_trading_index/st41_result.php?l=zh-tw&d={roc_y}/{m:02d}&o=json'
            r = requests.get(url, headers=headers, timeout=10)
            data = r.json()
            tables = data.get('tables', [])
            if tables:
                rows = tables[0].get('data', [])
                for row in rows:
                    # row[0] = '115/03/13', row[4] = 312.9 (櫃買指數)
                    parts = row[0].split('/')
                    if len(parts) == 3:
                        dt = f"{int(parts[0]) + 1911}-{parts[1]}-{parts[2]}"
                        val = row[4]
                        try:
                            tpex[dt] = float(str(val).replace(',', ''))
                        except ValueError:
                            pass
        except Exception as e:
            print(f"  ⚠️ TPEX {y}/{m:02d} 抓取失敗: {e}")

        # Next month
        m += 1
        if m > 12:
            m = 1
            y += 1

    print(f"  📈 TAIEX: {len(taiex)} 天, TPEX: {len(tpex)} 天")
    return taiex, tpex


# 抓取市場指數（快取到變數）
TAIEX = {}
TPEX = {}


def clean(obj):
    """清理 numpy/pandas 型別，確保 JSON 可序列化"""
    if isinstance(obj, dict):
        return {str(k): clean(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean(v) for v in obj]
    elif isinstance(obj, (np.integer,)):
        return int(obj)
    elif isinstance(obj, (np.floating, float)):
        if obj != obj:  # NaN
            return None
        return round(float(obj), 4)
    elif isinstance(obj, (np.bool_,)):
        return bool(obj)
    elif hasattr(obj, 'item'):
        return obj.item()
    elif hasattr(obj, 'isoformat'):
        return str(obj)
    return obj


# ═══════════════════════════════════════
# 數據載入函數
# ═══════════════════════════════════════

def _parse_981a_xlsx(fp, date_str):
    """解析 00981A 的 xlsx 檔案
    結構：
      Row 1:  資料日期
      Row 8:  項目 / 金額 / 權重 (期貨、股票)
      Row 12: 項目 / 金額 / 權重 (現金、保證金等)
      Row 20: 股票代號 / 股票名稱 / 股數 / 持股權重
      Row 21+: 持股資料
    """
    import openpyxl
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    def parse_pct(val):
        if val is None:
            return 0.0
        s = str(val).replace('%', '').replace(',', '').strip()
        try:
            return float(s)
        except ValueError:
            return 0.0

    # Parse cash & futures from rows with known labels
    cash_pct = 0
    futures_pct = 0
    for r in all_rows:
        if not r or not r[0]:
            continue
        label = str(r[0]).strip()
        if label == '現金' and len(r) >= 3:
            cash_pct = parse_pct(r[2])
        elif label == '期貨保證金' and len(r) >= 3:
            futures_pct = parse_pct(r[2])
        elif '期貨' in label and '名目' in label and len(r) >= 3:
            futures_pct = max(futures_pct, parse_pct(r[2]))

    # Find stock data start (after row with "股票代號")
    stock_start = None
    for i, r in enumerate(all_rows):
        if r and r[0] and str(r[0]).strip() == '股票代號':
            stock_start = i + 1
            break

    holdings = []
    if stock_start:
        for r in all_rows[stock_start:]:
            if not r or not r[0]:
                continue
            code = str(r[0]).strip()
            if not code or code == 'None':
                continue
            name = str(r[1]).strip() if len(r) > 1 and r[1] else ''
            # Weight is in column 4 (index 3) = '持股權重'
            weight = parse_pct(r[3]) if len(r) > 3 else 0
            if weight > 0:
                holdings.append({
                    'code': code,
                    'name': name,
                    'weight': round(weight, 2)
                })

    return {
        'holdings': holdings,
        'cash_pct': round(cash_pct, 2),
        'futures_pct': round(futures_pct, 2),
        'meta': {}
    }


def load_981a_data():
    """載入 00981A 所有日期的持股資料"""
    xlsx_dir = BASE / "00981A" / "daily_xlsx"
    records = {}

    # Old format: ETF_Investment_Portfolio_YYYYMMDD.xlsx
    for fp in sorted(glob.glob(str(xlsx_dir / "ETF_Investment_Portfolio_*.xlsx"))):
        fname = os.path.basename(fp)
        date_str = fname.replace("ETF_Investment_Portfolio_", "").replace(".xlsx", "")
        try:
            dt = datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")
        except:
            continue
        try:
            records[dt] = _parse_981a_xlsx(fp, dt)
        except Exception as e:
            print(f"  ⚠️ Skip {fname}: {e}")

    # New format: 00981A_YYYY-MM-DD.xlsx
    for fp in sorted(glob.glob(str(xlsx_dir / "00981A_*.xlsx"))):
        fname = os.path.basename(fp)
        date_str = fname.replace("00981A_", "").replace(".xlsx", "")
        try:
            datetime.strptime(date_str, "%Y-%m-%d")  # Validate format
            dt = date_str
        except:
            continue
        try:
            records[dt] = _parse_981a_xlsx(fp, dt)
        except Exception as e:
            print(f"  ⚠️ Skip {fname}: {e}")

    return records


def _parse_other_etf_xlsx(fp, etf_id):
    """解析 00980A/00982A/00991A/00993A 的 xlsx
    各檔格式略有不同，統一處理。
    """
    import openpyxl

    # 00991A 有時下載到的是 HTML（系統公告），跳過
    with open(fp, 'rb') as f:
        if f.read(4) != b'PK\x03\x04':
            return None

    wb = openpyxl.load_workbook(fp, data_only=True)

    def parse_pct(val):
        if val is None:
            return 0.0
        s = str(val).replace('%', '').replace(',', '').strip()
        try:
            return float(s)
        except ValueError:
            return 0.0

    holdings = []
    cash_pct = 0

    # 00982A: 持股在 '股票' sheet
    if '股票' in wb.sheetnames:
        ws = wb['股票']
        for row in range(2, ws.max_row + 1):
            code = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            weight = ws.cell(row=row, column=3).value
            if code and name and weight:
                w = parse_pct(weight)
                if w > 0:
                    holdings.append({'code': str(code).strip(), 'name': str(name).strip(), 'weight': round(w, 2)})
        # Cash from '其他資產' sheet
        if '其他資產' in wb.sheetnames and '投資組合' in wb.sheetnames:
            ws2 = wb['其他資產']
            nav_str = wb['投資組合'].cell(row=1, column=2).value or '0'
            nav = float(str(nav_str).replace('TWD', '').replace(',', '').replace('$', '').strip() or 0)
            for row in range(1, ws2.max_row + 1):
                label = str(ws2.cell(row=row, column=1).value or '').strip()
                if label == '現金':
                    cash_val = str(ws2.cell(row=row, column=2).value or '0')
                    cash_val = float(cash_val.replace('TWD', '').replace(',', '').replace('$', '').strip() or 0)
                    if nav > 0:
                        cash_pct = round(cash_val / nav * 100, 2)
    else:
        # 00980A / 00993A: 單 sheet，找 '股票代號' 行
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        stock_start = None
        weight_col = None  # 權重所在的 column index

        for i, r in enumerate(all_rows):
            if r and r[0] and str(r[0]).strip() == '股票代號':
                stock_start = i + 1
                # 找 '權重' 在第幾欄
                for j, cell in enumerate(r):
                    if cell and '權重' in str(cell):
                        weight_col = j
                        break
                break

        if stock_start and weight_col is not None:
            for r in all_rows[stock_start:]:
                if not r or not r[0]:
                    continue
                code = str(r[0]).strip()
                if not code or code == 'None':
                    continue
                name = str(r[1]).strip() if len(r) > 1 and r[1] else ''
                w = parse_pct(r[weight_col]) if len(r) > weight_col else 0
                if w > 0:
                    holdings.append({'code': code, 'name': name, 'weight': round(w, 2)})

        # Parse cash from meta rows
        for r in all_rows:
            if r and r[0] and str(r[0]).strip() == '現金' and len(r) >= 2:
                # Try to calculate cash % from NAV
                # For simplicity, estimate from stock weight sum
                break

    wb.close()

    if not holdings:
        return None

    stock_wt = sum(h['weight'] for h in holdings)
    if cash_pct == 0:
        cash_pct = round(max(0, 100 - stock_wt), 2)

    return {
        'holdings': sorted(holdings, key=lambda h: h['weight'], reverse=True),
        'cash_pct': cash_pct,
        'futures_pct': 0
    }


def load_other_etf(etf_id):
    """載入其他 ETF：優先用 daily_xlsx，再補 Master.csv 的舊資料"""
    records = {}

    # 1. 從 daily_xlsx 讀取（最新資料）
    xlsx_dir = BASE / etf_id / "daily_xlsx"
    if xlsx_dir.exists():
        for fp in sorted(xlsx_dir.glob(f"{etf_id}_*.xlsx")):
            date_str = fp.stem.split("_")[-1]
            if len(date_str) != 10:
                continue
            try:
                result = _parse_other_etf_xlsx(fp, etf_id)
                if result:
                    records[date_str] = result
            except Exception as e:
                print(f"  ⚠️ Skip {fp.name}: {e}")

    # 2. 從 Master.csv 補充更早的資料
    csv_path = BASE / etf_id / f"{etf_id}_Master.csv"
    if csv_path.exists():
        try:
            df = pd.read_csv(csv_path, encoding='utf-8-sig')
            col_map = {}
            for c in df.columns:
                cl = c.strip()
                if '代號' in cl: col_map[c] = 'code'
                elif '名稱' in cl: col_map[c] = 'name'
                elif '權重' in cl: col_map[c] = 'weight'
                elif '日期' in cl: col_map[c] = 'date'
            df = df.rename(columns=col_map)
            if 'date' in df.columns and 'weight' in df.columns:
                df['weight'] = pd.to_numeric(df['weight'], errors='coerce').fillna(0)
                for dt, group in df.groupby('date'):
                    dt_str = str(dt)
                    if dt_str in records:
                        continue  # xlsx 優先
                    holdings = []
                    for _, row in group.iterrows():
                        code = str(row.get('code', '')).strip()
                        name = str(row.get('name', '')).strip()
                        weight = float(row.get('weight', 0))
                        if code and name and weight > 0:
                            holdings.append({'code': code, 'name': name, 'weight': round(weight, 2)})

                    if holdings:
                        stock_wt = sum(h['weight'] for h in holdings)
                        cash_est = round(max(0, 100 - stock_wt), 2)
                        records[dt_str] = {
                            'holdings': sorted(holdings, key=lambda h: h['weight'], reverse=True),
                            'cash_pct': cash_est,
                            'futures_pct': 0
                        }
        except Exception as e:
            print(f"  ⚠️ Master.csv 讀取失敗: {e}")

    return records


# ═══════════════════════════════════════
# Signal Engine (精簡版)
# ═══════════════════════════════════════

def calc_cash_mode(records_981a, dates):
    """計算 00981A 攻防模式"""
    if not dates:
        return {}

    latest = dates[-1]
    rec = records_981a[latest]
    cash_now = rec.get('cash_pct') or 0
    futures_pct = rec.get('futures_pct') or 0
    n_holdings = len(rec.get('holdings', []))

    # Calculate 5MA and 20MA
    cash_vals = [records_981a[d].get('cash_pct', 0) or 0 for d in dates]
    cash_5ma = round(np.mean(cash_vals[-5:]), 2) if len(cash_vals) >= 5 else None
    cash_20ma = round(np.mean(cash_vals[-20:]), 2) if len(cash_vals) >= 20 else None

    # Determine mode
    if cash_now >= 7:
        mode = "🔴 高度防守"
        mode_desc = "現金水位高，經理人明確看空或避險"
    elif cash_now >= 5:
        mode = "🟡 防守"
        mode_desc = "現金偏高，保守觀望"
    elif cash_now >= 3:
        mode = "🟢 中性偏攻"
        mode_desc = "正常配置，略偏積極"
    else:
        mode = "🔵 積極進攻"
        mode_desc = "現金偏低，經理人積極做多"

    # Trend
    if len(cash_vals) >= 2:
        if cash_vals[-1] > cash_vals[-2] + 0.5:
            trend = "⬆️ 減倉中（現金增加）"
        elif cash_vals[-1] < cash_vals[-2] - 0.5:
            trend = "⬇️ 加倉中（現金減少）"
        else:
            trend = "➡️ 持平"
    else:
        trend = "—"

    # Build cash series
    cash_series = []
    for i, dt in enumerate(dates):
        r = records_981a[dt]
        cp = r.get('cash_pct', 0) or 0
        fp = r.get('futures_pct', 0) or 0
        sp = sum(h['weight'] for h in r.get('holdings', []))
        nh = len(r.get('holdings', []))

        c5 = round(np.mean([records_981a[dates[j]].get('cash_pct', 0) or 0 for j in range(max(0, i-4), i+1)]), 2) if i >= 4 else None
        c20 = round(np.mean([records_981a[dates[j]].get('cash_pct', 0) or 0 for j in range(max(0, i-19), i+1)]), 2) if i >= 19 else None

        cash_series.append({
            'date': dt, 'cash_pct': round(cp, 2), 'stock_pct': round(sp, 2),
            'futures_pct': round(fp, 2), 'n_holdings': nh,
            'cash_5ma': c5, 'cash_20ma': c20
        })

    has_futures = futures_pct > 0
    futures_signal = f"⚠️ 期貨保證金 {futures_pct}%，經理人可能進行避險或方向性操作" if has_futures else ""

    return {
        'mode': mode, 'mode_desc': mode_desc, 'trend': trend,
        'cash_now': round(cash_now, 2), 'cash_5ma': cash_5ma, 'cash_20ma': cash_20ma,
        'has_futures': has_futures, 'futures_signal': futures_signal,
        'n_holdings': n_holdings, 'cash_series': cash_series
    }


def calc_consensus(all_data, all_dates):
    """計算多 ETF 共識標的"""
    # Find the latest date for each ETF
    stock_map = {}
    for etf_id, records in all_data.items():
        etf_dates = sorted(records.keys())
        if not etf_dates:
            continue
        latest = etf_dates[-1]
        prev = etf_dates[-2] if len(etf_dates) >= 2 else None

        for h in records[latest].get('holdings', []):
            code = h['code']
            if code not in stock_map:
                stock_map[code] = {'name': h['name'], 'etf_weights': {}, 'changes': {}}
            stock_map[code]['etf_weights'][etf_id] = h['weight']

            # Calculate change
            if prev:
                prev_holdings = {hh['code']: hh['weight'] for hh in records[prev].get('holdings', [])}
                chg = h['weight'] - prev_holdings.get(code, 0)
                stock_map[code]['changes'][etf_id] = round(chg, 2)

    consensus = []
    for code, info in stock_map.items():
        n_etfs = len(info['etf_weights'])
        if n_etfs < 2:
            continue
        avg_weight = round(sum(info['etf_weights'].values()) / n_etfs, 2)
        total_weight = round(sum(info['etf_weights'].values()), 2)
        net_change = round(sum(info['changes'].values()), 2)
        n_adding = sum(1 for c in info['changes'].values() if c > 0.05)
        n_reducing = sum(1 for c in info['changes'].values() if c < -0.05)

        consensus.append({
            'code': code, 'name': info['name'], 'n_etfs': n_etfs,
            'etf_weights': info['etf_weights'], 'avg_weight': avg_weight,
            'total_weight': total_weight, 'net_change': net_change,
            'n_adding': n_adding, 'n_reducing': n_reducing,
            'etf_list': list(info['etf_weights'].keys())
        })

    consensus.sort(key=lambda c: c['total_weight'], reverse=True)
    return consensus[:20]


def calc_conviction(records_981a, dates, lookback=20):
    """計算信心度排行"""
    if len(dates) < 2:
        return []

    latest = dates[-1]
    start_idx = max(0, len(dates) - lookback)
    start_dt = dates[start_idx]

    current = {h['code']: h for h in records_981a[latest].get('holdings', [])}
    start_holdings = {h['code']: h for h in records_981a[start_dt].get('holdings', [])}

    result = []
    for code, h in current.items():
        start_w = start_holdings.get(code, {}).get('weight', 0)
        chg = round(h['weight'] - start_w, 2)

        if chg > 1:
            conv = "⭐⭐⭐ 高信心加碼"
        elif chg > 0.3:
            conv = "⭐⭐ 穩定加碼"
        elif chg > 0:
            conv = "⭐ 微幅增持"
        elif chg < -1:
            conv = "📉 大幅減碼"
        elif chg < -0.3:
            conv = "⬇️ 減碼中"
        else:
            conv = "➡️ 持平"

        result.append({
            'code': code, 'name': h['name'], 'weight': round(h['weight'], 2),
            'start_weight': round(start_w, 2), 'weight_chg': chg,
            'days': lookback, 'conviction': conv
        })

    result.sort(key=lambda r: abs(r['weight_chg']), reverse=True)
    return result[:20]


def calc_laomo_signals(records_981a, dates):
    """計算老墨跟單信號"""
    signals = []
    for i in range(1, len(dates)):
        dt = dates[i]
        prev_dt = dates[i-1]
        curr = {h['code']: h for h in records_981a[dt].get('holdings', [])}
        prev = {h['code']: h for h in records_981a[prev_dt].get('holdings', [])}

        for code, h in curr.items():
            if code not in prev:
                # New stock
                signals.append({
                    'date': dt, 'code': code, 'name': h['name'],
                    'type': '新增', 'weight': round(h['weight'], 2), 'weight_chg': round(h['weight'], 2),
                    'hold_suggestion': '觀察10日行情',
                    'confidence': '⭐⭐' if h['weight'] > 1 else '⭐'
                })
            else:
                chg = h['weight'] - prev[code]['weight']
                if chg > 0.3:
                    if chg > 1:
                        conf = '⭐⭐⭐'
                        hold = '持有60日，歷史勝率高'
                    elif chg > 0.5:
                        conf = '⭐⭐'
                        hold = '持有60日'
                    else:
                        conf = '⭐'
                        hold = '觀察20日'

                    signals.append({
                        'date': dt, 'code': code, 'name': h['name'],
                        'type': '加碼', 'weight': round(h['weight'], 2), 'weight_chg': round(chg, 2),
                        'hold_suggestion': hold, 'confidence': conf
                    })

    return signals


# ═══════════════════════════════════════
# 主生成函數
# ═══════════════════════════════════════

def generate():
    print(f"🐺 Wolf Pack Dashboard v5 — Data Generator")
    print(f"   Base: {BASE}")
    print(f"   Output: {OUTPUT_DIR}")
    print()

    # Load all ETF data
    print("📂 Loading 00981A xlsx...")
    records_981a = load_981a_data()
    dates_981a = sorted(records_981a.keys())
    print(f"   → {len(dates_981a)} dates: {dates_981a[0]} ~ {dates_981a[-1]}")

    all_data = {'00981A': records_981a}

    for etf_id in ['00980A', '00982A', '00991A', '00993A']:
        print(f"📂 Loading {etf_id} (xlsx + csv)...")
        records = load_other_etf(etf_id)
        if records:
            all_data[etf_id] = records
            etf_dates = sorted(records.keys())
            print(f"   → {len(etf_dates)} dates: {etf_dates[0]} ~ {etf_dates[-1]}")
        else:
            print(f"   → No data found")

    # Fetch market indices
    global TAIEX, TPEX
    print("\n📈 Fetching market indices (TAIEX + TPEX)...")
    TAIEX, TPEX = fetch_market_indices(2025, 10)

    # ═══════════════════════════════════════
    # Generate dashboard.json
    # ═══════════════════════════════════════
    print("\n⚙️ Generating dashboard.json...")

    v5 = {}
    v5['report_date'] = dates_981a[-1]
    v5['dates'] = dates_981a

    # Cash mode
    v5['cash_mode'] = calc_cash_mode(records_981a, dates_981a)

    # Cash series
    cs = []
    for i, dt in enumerate(dates_981a):
        rec = records_981a[dt]
        cash_pct = rec.get('cash_pct') or 0
        stock_pct = sum(h['weight'] for h in rec.get('holdings', []))
        n_holdings = len(rec.get('holdings', []))

        if i >= 4:
            vals = [cs[j]['cash_pct'] for j in range(i-4, i)] + [cash_pct]
            c5 = round(sum(vals)/5, 2)
        else:
            c5 = None
        if i >= 19:
            vals = [cs[j]['cash_pct'] for j in range(i-19, i)] + [cash_pct]
            c20 = round(sum(vals)/20, 2)
        else:
            c20 = None

        cs.append({'date': dt, 'cash_pct': round(cash_pct, 2), 'cash_5ma': c5, 'cash_20ma': c20,
                    'stock_pct': round(stock_pct, 2), 'n_holdings': n_holdings,
                    'taiex': TAIEX.get(dt), 'tpex': TPEX.get(dt)})
    v5['cash_series'] = cs

    # Latest holdings
    latest_holdings = {}
    for etf_id, records in all_data.items():
        etf_dates = sorted(records.keys())
        if not etf_dates:
            continue
        last_dt = etf_dates[-1]
        rec = records[last_dt]
        holdings = sorted(rec.get('holdings', []), key=lambda h: h['weight'], reverse=True)
        latest_holdings[etf_id] = {
            'date': last_dt, 'cash_pct': round(rec.get('cash_pct', 0) or 0, 2),
            'n_stocks': len(holdings),
            'stocks': [{'code': h['code'], 'name': h['name'], 'weight': round(h['weight'], 2)} for h in holdings]
        }
    v5['latest_holdings'] = latest_holdings

    # Weight history (top 15, 30 days)
    recent = dates_981a[-30:] if len(dates_981a) >= 30 else dates_981a
    top15 = latest_holdings.get('00981A', {}).get('stocks', [])[:15]
    wh = {}
    for s in top15:
        code = s['code']
        hist = []
        for dt in recent:
            w = 0
            for h in records_981a.get(dt, {}).get('holdings', []):
                if h['code'] == code:
                    w = round(h['weight'], 2)
                    break
            hist.append({'date': dt, 'weight': w})
        wh[code] = hist
    v5['weight_history'] = wh

    # Conviction, Consensus
    v5['conviction'] = calc_conviction(records_981a, dates_981a)
    v5['consensus'] = calc_consensus(all_data, dates_981a)

    # Daily changes (last 10)
    changes_list = []
    for i in range(max(0, len(dates_981a)-10), len(dates_981a)):
        dt = dates_981a[i]
        curr_map = {h['code']: h for h in records_981a[dt].get('holdings', [])}
        day = {'date': dt, 'new': [], 'added': [], 'reduced': [], 'exited': [],
               'cash_pct': cs[i]['cash_pct']}

        if i > 0:
            prev_map = {h['code']: h for h in records_981a[dates_981a[i-1]].get('holdings', [])}
            for code, h in curr_map.items():
                if code not in prev_map:
                    day['new'].append({'code': code, 'name': h['name'], 'weight': round(h['weight'], 2)})
                else:
                    chg = h['weight'] - prev_map[code]['weight']
                    if chg > 0.1:
                        day['added'].append({'code': code, 'name': h['name'], 'weight': round(h['weight'], 2), 'weight_chg': round(chg, 2)})
                    elif chg < -0.1:
                        day['reduced'].append({'code': code, 'name': h['name'], 'weight': round(h['weight'], 2), 'weight_chg': round(chg, 2)})
            for code, h in prev_map.items():
                if code not in curr_map:
                    day['exited'].append({'code': code, 'name': h['name']})
        changes_list.append(day)
    v5['daily_changes'] = {'00981A': changes_list}

    # Laomo signals
    v5['laomo_signals'] = calc_laomo_signals(records_981a, dates_981a)

    # Top 20 stocks
    v5['top20_stocks'] = latest_holdings.get('00981A', {}).get('stocks', [])[:20]

    # Stock series (top 20)
    stock_series = []
    for s in v5['top20_stocks']:
        code = s['code']
        weights = []
        for dt in dates_981a:
            w = 0
            for h in records_981a.get(dt, {}).get('holdings', []):
                if h['code'] == code:
                    w = round(h['weight'], 2)
                    break
            weights.append(w)
        stock_series.append({'code': code, 'label': f"{s['name']}({code})", 'data': weights})
    v5['stock_series'] = stock_series

    # Big add counts
    big_add = [0]
    for i in range(1, len(dates_981a)):
        curr = {h['code']: h['weight'] for h in records_981a[dates_981a[i]].get('holdings', [])}
        prev = {h['code']: h['weight'] for h in records_981a[dates_981a[i-1]].get('holdings', [])}
        count = sum(1 for c in curr if c in prev and curr[c] - prev[c] > 0.5)
        big_add.append(count)
    v5['big_add_counts'] = big_add

    # ═══════════════════════════════════════
    # Generate etf_pages.json
    # ═══════════════════════════════════════
    print("⚙️ Generating etf_pages.json...")

    etf_pages = {}
    for etf_id, records in all_data.items():
        etf_dates = sorted(records.keys())
        date_records = []
        cash_series_etf = []

        for dt in etf_dates:
            rec = records[dt]
            holdings = sorted(rec.get('holdings', []), key=lambda h: h['weight'], reverse=True)
            stock_wt = sum(h['weight'] for h in holdings)
            cash_pct = rec.get('cash_pct') or round(max(0, 100 - stock_wt), 2)
            taiex = TAIEX.get(dt)

            date_records.append({
                'date': dt,
                'holdings': [{'code': h['code'], 'name': h['name'], 'weight': round(h['weight'], 2)} for h in holdings],
                'cash_pct': round(cash_pct, 2),
                'n_stocks': len(holdings),
                'stock_weight': round(stock_wt, 2),
                'taiex': taiex
            })
            cash_series_etf.append({'date': dt, 'cash_pct': round(cash_pct, 2), 'taiex': taiex, 'tpex': TPEX.get(dt)})

        etf_pages[etf_id] = {
            'dates': etf_dates, 'n_dates': len(etf_dates),
            'date_records': date_records, 'cash_series': cash_series_etf
        }

    # ═══════════════════════════════════════
    # Save
    # ═══════════════════════════════════════
    v5 = clean(v5)
    etf_pages = clean(etf_pages)

    with open(OUTPUT_DIR / 'dashboard.json', 'w', encoding='utf-8') as f:
        json.dump(v5, f, ensure_ascii=False)

    with open(OUTPUT_DIR / 'etf_pages.json', 'w', encoding='utf-8') as f:
        json.dump(etf_pages, f, ensure_ascii=False)

    sz1 = (OUTPUT_DIR / 'dashboard.json').stat().st_size
    sz2 = (OUTPUT_DIR / 'etf_pages.json').stat().st_size
    print(f"\n✅ dashboard.json: {sz1:,} bytes ({sz1/1024:.0f} KB)")
    print(f"✅ etf_pages.json: {sz2:,} bytes ({sz2/1024:.0f} KB)")
    print(f"📅 Report date: {v5['report_date']}")
    print(f"🐺 Done!")


if __name__ == '__main__':
    generate()
